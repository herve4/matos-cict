import datetime
import random
import sqlite3

import barcode
import easyocr
import numpy as np
import pandas as pd

import streamlit as st
from barcode.writer import ImageWriter
from openpyxl.styles import Font
from openpyxl.utils import get_column_letter
from openpyxl.workbook import Workbook


def connexion():
    db = sqlite3.connect("materiels_db")
    sql = (
        'CREATE TABLE IF NOT EXISTS materiels (id_materiels INT(15) PRIMARY KEY,designation VARCHAR(100),prix VARCHAR(10),fournisseurs '
        'VARCHAR(100), date VARCHAR(8), service VARCHAR(80),image VARCHAR(255),codeB VARCHAR(255), codeBarText VARCHAR(20));')
    cursor = db.cursor()
    cursor.execute(sql)


connexion()
title = st.set_page_config(page_title="Matériels CICT", layout="wide", page_icon="Vb8grj_d_400x400.ico")
with st.container():
    st.image(image="64ae983494fa9-materiaux de construction au Bénin.jpg",use_column_width="auto")
col1, col2 = st.columns(2)


class My_app(object):
    def __init__(self):
        super(My_app, self).__init__()

        self.id = random.randrange(1000, 9999)
        self.data = dict()


        self.search = col2.text_input("Rechercher par ID ou scanner un article",
                                      placeholder="1000 ou 1233434509815")

        list_mat = col2.button("Voir la liste des matériels")
        list_scan = col2.button("Voir le matériel scanné")
        if list_mat:
            self.My_tab()

        with st.container(border=True):
            if list_scan:
                self.scanners()

        self.imputs()

    def scanners(self):
       if self.list_scann() is not None:
           m = []
           for item in self.list_scann():
               m.append(item)
           with col2.container(border=True):
               df = pd.DataFrame(self.list_scann(),columns=["Id","Code barre","Date"])
               st.title("Matériels scannés")

               st.table(df)


       else:
           col2.warning("Aucun scanne effectué pour cet article !")

    def loadList(self):

        db = sqlite3.connect("materiels_db")
        cursor = db.cursor()
        sql = "SELECT id_materiels,codeBarText,designation,prix,fournisseurs,date,service,codeB FROM materiels"
        self.result = cursor.execute(sql)
        return self.result

    def My_tab(self):
        m = []
        for item in self.loadList():
            m.append(item)
        df = pd.DataFrame(self.loadList(),
                          columns=["ID", "Code", "Désignation", "Prix", "Fournisseurs", "Date d'arrivée", "Service",
                                   "Image"])
        col2.title("Liste de matériels")

        # self.col2.dataframe(df)

        column_configuration = {
            "text": st.column_config.TextColumn(
                "Désignation",
            ),
            "price": st.column_config.NumberColumn(
                "Prix",
                min_value=0,
                max_value=120,
                help="The user's age",
            ),
            "image": st.column_config.ImageColumn("Image", help="Image de l'article", width="small"),

            "date": st.column_config.DateColumn(
                "Date d'arrivée",
                help="Date où l'article est arrivé dans l'entreprise",
                min_value=datetime.date(1920, 1, 1),
                format="%d years",
            ),
        }
        col2.data_editor(
            df,
            column_config=column_configuration,
            use_container_width=True,
            hide_index=True,
            num_rows="fixed",
        )

    def NumberGenerter(self):
        num1 = "0123456789"
        num2 = "0123456789"
        number = num1 + num2
        length = 13
        result = "".join(random.sample(number, length))
        return result

    def imputs(self):
        with col1.form("My_form", clear_on_submit=True):
            self.header = st.header("Ajouter un article")
            self.des = st.text_input("Désignation")
            self.prix = st.number_input("Prix")
            self.frns = st.text_input("Fournisseurs")
            self.date = st.date_input("Date")
            self.service = st.text_input("Service")
            self.image_up = st.file_uploader("Importer une image", key=6)
            self.btn = st.form_submit_button("Enregistrer", type="secondary")
            if self.btn:
                self.barcode_Generate()

    def barcode_Generate(self):

        if len(self.des) > 0:
            if self.prix is not None:
                if len(self.frns) > 0:
                    if len(self.service) > 0:
                        if self.date is not None:
                            if self.image_up is not None:
                                stringio = self.image_up.getvalue()

                                self.generatedBarcode = []
                                # set the barcode format
                                barcode_format = barcode.get_barcode_class('ean')
                                # Barcode Number
                                self.barcodeNumber = self.NumberGenerter()  # 12 digit
                                generated = barcode_format(self.barcodeNumber, writer=ImageWriter())

                                generated.save(self.des)
                                self.generatedBarcode.append(f"{self.des}.png")
                                self.data[self.barcodeNumber] = [f"{self.des}.png"]
                                image = self.data[self.barcodeNumber]

                                path_file = r"C:\Users\Default User\Downloads"

                                print("".join(image))
                                self.code_generate_finish = "".join(image)

                                read = easyocr.Reader(['fr'])
                                img_text = read.readtext("".join(image), detail=0)
                                f = []
                                for t in img_text:
                                    f.append(t)
                                    print(t)
                                print(int("".join(f)))
                                db = sqlite3.connect("materiels_db")
                                c = db.cursor()
                                q = f"INSERT INTO materiels(id_materiels,designation,prix,fournisseurs,date,service,image,codeB,codeBarText) VALUES(?,?,?,?,?,?,?,?,?)"
                                c.execute(q, (self.id,
                                              self.des,
                                              self.prix,
                                              self.frns,
                                              self.date,
                                              self.service,
                                              stringio,
                                              "".join(image),
                                              int("".join(f))
                                              ))

                                db.commit()
                                db.close()

                                col2.image(self.image_up)
                                col2.image("".join(image))
                                col2.download_button(label="Enregistrer le code barre", data="".join(image),
                                                     file_name=f"{self.des}")
                                col1.success("L'article a été ajouté avec succès!")


                        else:
                            col1.warning("Selectionner une date de livraison ou d'achat !")
                    else:
                        col1.warning("Veuillez definir le service !")
                else:
                    col1.warning("Vous devez insérer un fournisseur !")
            else:
                col1.warning("Insérer un prix !")
        else:
            col1.warning("Veuillez insérer une désignation")

    def list_scann(self):
        r = "SELECT id_scan,codeBarText,date FROM scan ;"
        # row = (code_text,)
        My_code_bar = self.search
        db = sqlite3.connect("materiels_db")
        c = db.cursor()
        c.execute(r)
        result = c.fetchall()
        return result
    def windows_play_code_bar_select(self, code):
        r = f"SELECT * FROM materiels WHERE codeBarText=?;"
        # row = (code_text,)
        My_code_bar = code
        db = sqlite3.connect("materiels_db")
        c = db.cursor()
        c.execute(r, (My_code_bar,))
        result = c.fetchone()

        return result

    def add_code_and_display(self):

        result = self.windows_play_code_bar_select(self.search)
        # print(result)
        date_joined = str(datetime.datetime.now())
        print(date_joined)
        if result:

            # col2.write(result[1])
            # col2.write(result[4])
            # col2.write(f"{result[2]} F CFA")
            # col2.write(result[3])
            # col2.write(result[5])
            # col2.image(result[6])
            results = [result[1], result[4], result[2], result[3], result[5], date_joined, result[7]]
            # self.code_bar_img.setText(f"- Code barre d'article : {result[8]}")
            print(results)

            data = {
                result[0]: {
                    "Désignation": result[1],
                    "Code barre": result[8],
                    "Prix": result[2],
                    "Fournisseurs": result[3],
                    "Service": result[5],
                    "Date d'arrivée": result[4],
                    "Scanné le": date_joined,
                    "Image": result[7]
                }
            }
            r = []
            for i in result:
                r.append(i)
            print(r)
            df = pd.DataFrame(r)
            df.to_excel("materiels.xlsx")

            wb = Workbook()
            ws = wb.active
            ws.title = "Liste Matériels C.I.C.T"

            headings = ['ID de l\'article'] + list(data[result[0]].keys())
            ws.append(headings)

            for person in data:
                grades = list(data[person].values())
                ws.append([person] + grades)

            for col in range(2, len(data[result[0]]) + 2):
                char = get_column_letter(col)
                ws[char + "7"] = f"=SUM({char + '2'}:{char + '6'})/{len(data)}"

            for col in range(1, 9):
                ws[get_column_letter(col) + '1'].font = Font(bold=True, color="0099CCFF")

            wb.save("Matériels.xlsx")
        else:
            col2.warning("Aucune info trouvé pour ce code !")

    def delete_customer(self, ID):
        conn = sqlite3.connect('materiels_db')
        c = conn.cursor()
        c.execute("DELETE FROM materiels WHERE id_materiels=?", (ID,))
        conn.commit()
        conn.close()

    def update_customer(self, des, frns, service, prix, date, code):
        conn = sqlite3.connect('materiels_db')
        c = conn.cursor()
        c.execute(
            "UPDATE materiels SET designation = ?,fournisseurs = ? , service = ?,prix=?, date=? WHERE codeBarText = ?",
            (des, frns, service, prix, date, code))
        conn.commit()
        conn.close()

    def search_customer(self, ID):

        if len(self.search) == 4:
            conn = sqlite3.connect('materiels_db')
            c = conn.cursor()
            c.execute(
                "SELECT id_materiels,codeBarText,designation,prix,fournisseurs,date,service FROM materiels WHERE Id_materiels=?",
                (ID,))
            article = c.fetchall()
            conn.close()
            return article
        elif len(self.search) == 13:
            conn = sqlite3.connect('materiels_db')
            c = conn.cursor()
            c.execute(
                "SELECT id_materiels,codeBarText,designation,prix,fournisseurs,date,service,image,codeB FROM materiels WHERE codeBarText=?",
                (ID,))
            article = c.fetchone()

            conn.close()
            return article

    def search_affcihe_result(self):
        if len(self.search) == 4:
            articles = self.search_customer(self.search)
            df = pd.DataFrame(articles,
                              columns=["ID", "Code", "Désignation", "Prix", "Fournisseurs", "Date d'arrivée",
                                       "Service"])
            col2.dataframe(df)
        elif len(self.search) == 13:
            articles = self.search_customer(self.search)
            db = sqlite3.connect("materiels_db")
            sql = (
                'CREATE TABLE IF NOT EXISTS scan (id_scan INT(20) PRIMARY KEY,codeBarText VARCHAR(20),date VARCHAR(20));')
            cursor = db.cursor()
            cursor.execute(sql)

            sql2 = 'SELECT * FROM scan WHERE codeBarText=? ;'
            if not cursor.execute(sql2, (articles[1],)):

                # Inserrer les données scannées
    
                sql_scan = 'INSERT INTO scan (id_scan,codeBarText,date) VALUES(?,?,?);'
               
                 
                date_joined = str(datetime.datetime.now())
                f = db.cursor()
                f.execute(sql_scan, (articles[0], articles[1], date_joined))
            
            

                db.commit()
                db.close()
            if articles:

                l = []
                for item in articles:
                    l.append(item)
                print(l)
                with st.container(border=True):
                    col2.markdown(f"##### {articles[2]}")
                    col2.image(articles[7])
                    col2.write(f":red[ID] : {articles[0]}")
                    col2.write(f":red[Code barre] : {articles[1]}")
                    col2.write(f":red[Prix de l'article] : {articles[3]} F CFA")
                    col2.write(f":red[Fournisseurs] : {articles[4]}")
                    col2.write(f":red[Service] : {articles[6]}")
                    col2.write(f":red[Date d'arrivée] : {articles[5]}")
                    col2.image(articles[8])
                    self.add_code_and_display()




                    self.modif = col2.container(border=True)
                    with self.modif:
                        btn_modif = st.button("Modfifier cet article")
                        btn_suppr = st.button("Supprimer cet article",type="primary")
                        if btn_suppr:
                            self.delete_customer(articles[0])
                    if self.modif:
                        header = col1.header("Modification de l'article")
                        des = col1.text_input("Désignation", value=articles[2])
                        prix = col1.number_input("Prix", key=2)
                        frns = col1.text_input("Fournisseurs", value=articles[4])
                        date = col1.date_input("Date", key=3)
                        service = col1.text_input("Service", value=articles[6])
                        image_up = col1.file_uploader("Importer une image")
                        btn = col1.button("Modifier", type="secondary", key=5)
                        if btn:
                            self.update_customer(des, frns, service, prix, date, articles[1])
                            st.success("L'article a bien été modfier !")
            else:
                col2.warning("Aucune article trouvé pour ce code !")


if '__main__' == __name__:
    app = My_app()
    app.search_affcihe_result()
